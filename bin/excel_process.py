import os.path
from openpyxl import load_workbook

class WordDataset:
	def __init__(self):
		self.list_kata=[]
		tipe = []
		sense = []
		sentence = []
		self.read_dataset = read_dataset =load_workbook(filename = os.path.dirname(__file__)+'/../dataset/ambiguous word.xlsx').active
		max_row = read_dataset.max_row
		# max_column = read_dataset.max_column
		
		for row in range (4, max_row):
			if read_dataset["B"+str(row)].value != None:
				if row > 4:
					create_object = {
					"name": name,
					"tipe": tipe,
					"sense": sense,
					"sentence": sentence
					}
					self.list_kata.append(create_object)
				tipe = []
				sense = []
				sentence = []
				name = read_dataset["B"+str(row)].value
				tipe.append(read_dataset["C"+str(row)].value) 
				sense.append(read_dataset["D"+str(row)].value) 
				sentence.append(read_dataset["E"+str(row)].value)
			
			else :
				tipe.append(read_dataset["C"+str(row)].value) 
				sense.append(read_dataset["D"+str(row)].value) 
				sentence.append(read_dataset["E"+str(row)].value)
				
				
		create_object = {
			"name": name,
			"tipe": tipe,
			"sense": sense,
			"sentence": sentence
			}
		self.list_kata.append(create_object)

	def getListKata(self):
			return self.list_kata

